#!/usr/bin/env python3
"""
Warehouse Dashboard Auto-Updater
Runs via GitHub Actions to automatically update dashboard data
"""

import json
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os

def extract_health_tracker():
    """Extract data from Health Tracker - Row 36 and Row 34"""
    try:
        wb = load_workbook('Health_Tracker_2026_xlsx.xlsx', data_only=True)
        feb_sheet = wb['Feb']
        
        # Row 36 = Latest Day (if formulas exist)
        # Row 34 = Totals
        
        total_loaded = feb_sheet.cell(34, 8).value  # H34
        total_loaded_vs_drop = feb_sheet.cell(34, 13).value  # M34
        total_late_loads = feb_sheet.cell(34, 16).value  # P34
        total_roll_over = feb_sheet.cell(34, 17).value  # Q34
        
        # Count days
        day_count = 0
        for row_idx in range(6, 34):
            val = feb_sheet.cell(row_idx, 8).value
            cell_b = feb_sheet.cell(row_idx, 2).value
            if cell_b and 'total' in str(cell_b).lower():
                continue
            if val and val != 0:
                day_count += 1
        
        # Calculate bin statistics
        bin_values = []
        for row_idx in range(6, 34):
            cell_b = feb_sheet.cell(row_idx, 2).value
            if cell_b and 'total' in str(cell_b).lower():
                continue
            bins = feb_sheet.cell(row_idx, 19).value
            if bins and isinstance(bins, (int, float)) and bins > 0:
                bin_values.append(int(bins))
        
        avg_bins = int(sum(bin_values) / len(bin_values)) if bin_values else 0
        max_bins = max(bin_values) if bin_values else 0
        min_bins = min(bin_values) if bin_values else 0
        
        # Calculate averages
        avg_per_day = int(total_loaded / day_count) if day_count > 0 and total_loaded else 0
        avg_loaded_vs_drop = int(total_loaded_vs_drop / day_count) if day_count > 0 and total_loaded_vs_drop else 0
        avg_roll_over = round(total_roll_over / day_count, 1) if day_count > 0 and total_roll_over else 0
        
        return {
            'running_total': int(total_loaded) if total_loaded else 0,
            'days_tracked': day_count,
            'avg_per_day': avg_per_day,
            'loaded_vs_drop_total': int(total_loaded_vs_drop) if total_loaded_vs_drop else 0,
            'avg_loaded_vs_drop': avg_loaded_vs_drop,
            'late_loads': int(total_late_loads) if total_late_loads else 0,
            'roll_over': int(total_roll_over) if total_roll_over else 0,
            'avg_roll_over': avg_roll_over,
            'avg_bins': avg_bins,
            'max_bins': max_bins,
            'min_bins': min_bins,
        }
        
    except Exception as e:
        print(f"Error reading Health Tracker: {e}")
        return None

def extract_picker_data():
    """Extract picker efficiency data"""
    try:
        wb = load_workbook('Picker_Efficiency_2026_xlsx.xlsx')
        sheet = wb['Formulas']
        
        picker_data = []
        total_picked = 0
        total_hours = 0
        
        for row in sheet.iter_rows(min_row=3, max_row=30, values_only=True):
            if row[0] and row[1]:
                hours = row[2] if row[2] else 0
                pallets = sum([row[i] for i in range(3, 10) if row[i] and isinstance(row[i], (int, float))])
                if pallets > 0 or hours > 0:
                    total_picked += pallets
                    total_hours += hours
                    picker_data.append({
                        'name': f"{row[0]} {row[1]}",
                        'pallets': pallets,
                        'hours': hours,
                        'efficiency': round(pallets / hours, 2) if hours > 0 else 0
                    })
        
        return {
            'total': int(total_picked),
            'hours': round(total_hours, 2),
            'efficiency': round(total_picked / total_hours, 2) if total_hours > 0 else 0,
            'count': len(picker_data),
            'top_performers': sorted(picker_data, key=lambda x: x['efficiency'], reverse=True)[:5]
        }
        
    except Exception as e:
        print(f"Error reading Picker data: {e}")
        return None

def extract_putaway_data():
    """Extract putaway efficiency data"""
    try:
        wb = load_workbook('Putaway_Efficiency_2026_xlsx.xlsx')
        sheet = wb['Formulas']
        
        putaway_data = []
        total_putaway = 0
        total_hours = 0
        
        for row in sheet.iter_rows(min_row=3, max_row=30, values_only=True):
            if row[0] and row[1]:
                hours = row[2] if row[2] else 0
                pallets = sum([row[i] for i in range(3, 10) if row[i] and isinstance(row[i], (int, float))])
                if pallets > 0 or hours > 0:
                    total_putaway += pallets
                    total_hours += hours
                    putaway_data.append({
                        'name': f"{row[0]} {row[1]}",
                        'pallets': pallets,
                        'hours': hours,
                        'efficiency': round(pallets / hours, 2) if hours > 0 else 0
                    })
        
        return {
            'total': int(total_putaway),
            'hours': round(total_hours, 2),
            'efficiency': round(total_putaway / total_hours, 2) if total_hours > 0 else 0,
            'count': len(putaway_data),
            'top_performers': sorted(putaway_data, key=lambda x: x['efficiency'], reverse=True)[:5]
        }
        
    except Exception as e:
        print(f"Error reading Putaway data: {e}")
        return None

def main():
    print("=" * 60)
    print("WAREHOUSE DASHBOARD AUTO-UPDATE")
    print("=" * 60)
    print(f"Running at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Extract all data
    health_data = extract_health_tracker()
    picker_data = extract_picker_data()
    putaway_data = extract_putaway_data()
    
    if not health_data or not picker_data or not putaway_data:
        print("❌ ERROR: Could not extract all required data")
        return False
    
    # Create dashboard data structure
    dashboard_data = {
        'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'totals': {
            'loader_running_total': health_data['running_total'],
            'loader_days_tracked': health_data['days_tracked'],
            'loader_avg_per_day': health_data['avg_per_day'],
            'loader_date': datetime.now().strftime('%B %d, %Y'),
            'loader_date_short': datetime.now().strftime('%b %d'),
            'loader_month': 'February 2026',
            'loaded_vs_drop': health_data['loaded_vs_drop_total'],
            'avg_loaded_vs_drop': health_data['avg_loaded_vs_drop'],
            'late_loads': health_data['late_loads'],
            'roll_over': health_data['roll_over'],
            'avg_roll_over': health_data['avg_roll_over'],
            'avg_bins': health_data['avg_bins'],
            'max_bins': health_data['max_bins'],
            'min_bins': health_data['min_bins'],
            'loader_count': 28,
            'pallets_picked': picker_data['total'],
            'picker_hours': picker_data['hours'],
            'picker_efficiency': picker_data['efficiency'],
            'pallets_putaway': putaway_data['total'],
            'putaway_hours': putaway_data['hours'],
            'putaway_efficiency': putaway_data['efficiency'],
            'picker_count': picker_data['count'],
            'putaway_count': putaway_data['count'],
            'total_staff': 28 + picker_data['count'] + putaway_data['count'],
            'total_hours': round(picker_data['hours'] + putaway_data['hours'], 2),
        },
        'top_pickers': picker_data['top_performers'],
        'top_putaway': putaway_data['top_performers'],
    }
    
    # Save to JSON
    with open('warehouse_data.json', 'w') as f:
        json.dump(dashboard_data, f, indent=2)
    
    print("\n✅ SUCCESS!")
    print(f"   Loader Total: {dashboard_data['totals']['loader_running_total']:,}")
    print(f"   Days: {dashboard_data['totals']['loader_days_tracked']}")
    print(f"   Picker Total: {dashboard_data['totals']['pallets_picked']:,}")
    print(f"   Putaway Total: {dashboard_data['totals']['pallets_putaway']:,}")
    print(f"\n✓ warehouse_data.json updated")
    print("=" * 60)
    
    return True

if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)
